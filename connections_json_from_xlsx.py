#!/usr/bin/env python3

import argparse
import json
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_INPUT_FILE = Path("data/connections.xlsx")
DEFAULT_OUTPUT_FILE = Path("data/connections.json")
CONNECTIONS_SHEET = "Connections"
MATRIX_SHEET = "London Matrix"
DEFINITIONS_SHEET = "Definitions"
DEFAULT_MATRIX_MODE = "Underground"


def base_location(name):
    return re.sub(r"\([^)]*\)$", "", str(name).strip())


def station_list(value):
    if value is None:
        return None
    parts = [part.strip() for part in str(value).split(",")]
    parts = [part for part in parts if part]
    return parts or None


def parse_matrix_duration(value):
    if value is None:
        return None

    text = str(value).strip()
    if not text or text.upper() == "X":
        return None

    mode = DEFAULT_MATRIX_MODE
    if text.upper().endswith("W"):
        mode = "walk"
        text = text[:-1].strip()

    try:
        minutes = int(float(text))
    except ValueError:
        return None

    if minutes <= 0:
        return None

    return {
        "durationMinutes": minutes,
        "mode": mode,
    }


def parse_minutes(value, *, context):
    if value is None or str(value).strip() == "":
        raise ValueError(f"Missing minutes in {context}")
    try:
        minutes = int(float(str(value).strip()))
    except ValueError as exc:
        raise ValueError(f"Invalid minutes {value!r} in {context}") from exc
    if minutes <= 0:
        raise ValueError(f"Minutes must be positive in {context}: {value!r}")
    return minutes


def parse_mode(value, *, context):
    if value is None or str(value).strip() == "":
        raise ValueError(f"Missing mode in {context}")
    return str(value).strip()


def add_direction_fields(target, definition):
    previous = definition.get("previousStations")
    next_ = definition.get("nextStations")

    if previous:
        target["previousStations"] = previous
    if next_:
        target["nextStations"] = next_


def find_sheet(wb, preferred_name, fallback_names=()):
    names_by_lower = {name.lower(): name for name in wb.sheetnames}
    candidates = (preferred_name, *fallback_names)
    for candidate in candidates:
        actual = names_by_lower.get(candidate.lower())
        if actual:
            return wb[actual]
    raise KeyError(
        f"Could not find sheet {preferred_name!r}. Available sheets: {', '.join(wb.sheetnames)}"
    )


def read_definitions(wb):
    definitions_ws = find_sheet(wb, DEFINITIONS_SHEET, ("definitions",))
    definitions = {}

    for row_number, row in enumerate(definitions_ws.iter_rows(min_row=2, values_only=True), start=2):
        location, previous, next_ = row[:3]
        if location is None or str(location).strip() == "":
            continue

        definitions[str(location).strip()] = {
            "previousStations": station_list(previous),
            "nextStations": station_list(next_),
        }

    return definitions


def append_connection(output, origin_base, origin_entry, target_base, connection_entry):
    # Merge into an existing identical origin entry where possible. This matters for simple
    # CRS-to-CRS connections, which have no previous/next filters and therefore do not need
    # multiple direction-specific origin entries.
    origin_entries = output.setdefault(origin_base, [])

    matching_origin = None
    for existing_origin in origin_entries:
        existing_filters = {
            key: existing_origin.get(key)
            for key in ("previousStations", "nextStations")
            if key in existing_origin
        }
        new_filters = {
            key: origin_entry.get(key)
            for key in ("previousStations", "nextStations")
            if key in origin_entry
        }
        if existing_filters == new_filters:
            matching_origin = existing_origin
            break

    if matching_origin is None:
        matching_origin = dict(origin_entry)
        matching_origin["connections"] = {}
        origin_entries.append(matching_origin)

    target_entries = matching_origin.setdefault("connections", {}).setdefault(target_base, [])
    if connection_entry not in target_entries:
        target_entries.append(connection_entry)


def add_matrix_connections(wb, definitions, output):
    matrix_ws = find_sheet(wb, MATRIX_SHEET, ("matrix",))
    headers = [
        cell.value if cell.value is None else str(cell.value).strip()
        for cell in matrix_ws[1][1:]
    ]

    for row_number, row in enumerate(matrix_ws.iter_rows(min_row=2, values_only=True), start=2):
        origin = row[0]
        if origin is None or str(origin).strip() == "":
            continue

        origin = str(origin).strip()
        origin_base = base_location(origin)

        origin_entry = {}
        add_direction_fields(origin_entry, definitions.get(origin, {}))

        for target, raw_duration in zip(headers, row[1:]):
            if target is None or str(target).strip() == "":
                continue

            parsed_duration = parse_matrix_duration(raw_duration)
            if parsed_duration is None:
                continue

            target = str(target).strip()
            target_base = base_location(target)

            connection_entry = {}
            add_direction_fields(connection_entry, definitions.get(target, {}))
            connection_entry.update(parsed_duration)

            append_connection(output, origin_base, origin_entry, target_base, connection_entry)


def normalise_header(value):
    return re.sub(r"[^a-z0-9]+", "", str(value).lower()) if value is not None else ""


def add_simple_crs_connections(wb, output):
    connections_ws = find_sheet(wb, CONNECTIONS_SHEET, ("connections",))

    header_cells = next(connections_ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header_index = {normalise_header(value): idx for idx, value in enumerate(header_cells)}

    required = {
        "location1crs": "location 1 CRS",
        "location2crs": "location 2 CRS",
        "mins": "mins",
        "mode": "mode",
    }
    missing = [display for key, display in required.items() if key not in header_index]
    if missing:
        raise ValueError(f"Connections sheet is missing required column(s): {', '.join(missing)}")

    for row_number, row in enumerate(connections_ws.iter_rows(min_row=2, values_only=True), start=2):
        crs_1 = row[header_index["location1crs"]]
        crs_2 = row[header_index["location2crs"]]
        mins = row[header_index["mins"]]
        mode = row[header_index["mode"]]

        if all(value is None or str(value).strip() == "" for value in (crs_1, crs_2, mins, mode)):
            continue

        context = f"Connections row {row_number}"
        if crs_1 is None or str(crs_1).strip() == "":
            raise ValueError(f"Missing location 1 CRS in {context}")
        if crs_2 is None or str(crs_2).strip() == "":
            raise ValueError(f"Missing location 2 CRS in {context}")

        crs_1 = str(crs_1).strip().upper()
        crs_2 = str(crs_2).strip().upper()
        minutes = parse_minutes(mins, context=context)
        parsed_mode = parse_mode(mode, context=context)

        connection_entry = {
            "durationMinutes": minutes,
            "mode": parsed_mode,
        }

        # Simple CRS-to-CRS links are deliberately unfiltered: no previousStations or nextStations.
        # Add both directions so either station can be used as the origin.
        append_connection(output, crs_1, {}, crs_2, connection_entry)
        append_connection(output, crs_2, {}, crs_1, connection_entry)


def build_json(input_file):
    wb = load_workbook(input_file, data_only=True)
    definitions = read_definitions(wb)
    output = {}

    add_matrix_connections(wb, definitions, output)
    add_simple_crs_connections(wb, output)

    return output


def parse_args():
    parser = argparse.ArgumentParser(
        description="Export connections.xlsx to connections.json."
    )
    parser.add_argument(
        "input_file",
        nargs="?",
        type=Path,
        default=DEFAULT_INPUT_FILE,
        help=f"Input workbook path, default: {DEFAULT_INPUT_FILE}",
    )
    parser.add_argument(
        "output_file",
        nargs="?",
        type=Path,
        default=DEFAULT_OUTPUT_FILE,
        help=f"Output JSON path, default: {DEFAULT_OUTPUT_FILE}",
    )
    return parser.parse_args()


def main():
    args = parse_args()
    output = build_json(args.input_file)

    args.output_file.parent.mkdir(parents=True, exist_ok=True)
    with args.output_file.open("w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    print(f"Wrote {args.output_file}")


if __name__ == "__main__":
    main()
