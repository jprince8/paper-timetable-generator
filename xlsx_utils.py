import io
import re


INVALID_SHEET_CHARS = re.compile(r"[:\\/?*\[\]]")


def _sheet_name(name, fallback):
    text = INVALID_SHEET_CHARS.sub(" ", str(name or "").strip())[:31].strip()
    return text or fallback


def _cell_text(value):
    if value is None:
        return ""
    if isinstance(value, dict):
        return str(value.get("text") or "")
    return str(value)


def _append_rows(ws, rows):
    extra_rows = []
    for row in rows or []:
        _append_row(ws, row)
        label = _cell_text(row[0] if row else "").strip()
        if label in {"Comes from", "Continues to"}:
            extra_rows.append(ws.max_row)
    return extra_rows


def _append_row(ws, row):
    ws.append([_cell_text(cell) for cell in row])
    row_idx = ws.max_row
    for col_idx, value in enumerate(row, start=1):
        if not isinstance(value, dict):
            continue
        hyperlink = str(value.get("hyperlink") or "").strip()
        if not hyperlink:
            continue
        cell = ws.cell(row_idx, col_idx)
        cell.hyperlink = hyperlink
        cell.style = "Hyperlink"


def _fit_columns(ws, max_width=34, ignore_rows=None):
    from openpyxl.utils import get_column_letter

    ignore_rows = set(ignore_rows or [])
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        width = 10
        for cell in column_cells:
            if cell.row in ignore_rows:
                continue
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(max_width, len(value) + 2))
        ws.column_dimensions[letter].width = width


def _format_timetable_sheet(
    ws,
    title_rows,
    headcode_row,
    header_row,
    facility_row,
    extra_rows=None,
):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    max_row = ws.max_row
    max_col = ws.max_column
    thin_side = Side(style="thin", color="B8B8B8")
    medium_side = Side(style="medium", color="707070")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    station_top_border = Border(
        left=thin_side,
        right=thin_side,
        top=medium_side,
        bottom=thin_side,
    )
    title_font = Font(bold=True, size=14)
    date_font = Font(italic=True, color="666666")
    bold_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    facility_fill = PatternFill("solid", fgColor="EDEDED")
    band_fills = [
        PatternFill("solid", fgColor="FFFFFF"),
        PatternFill("solid", fgColor="F7FBFF"),
    ]

    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.border = border
            if col_idx == 1:
                horizontal = "left"
            elif col_idx == 2:
                horizontal = "right"
            else:
                horizontal = "center"
            cell.alignment = Alignment(
                horizontal=horizontal,
                vertical="center",
                wrap_text=True,
            )

    if title_rows:
        ws.cell(1, 1).font = title_font
        if max_col > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    if len(title_rows) > 1:
        ws.cell(2, 1).font = date_font
        if max_col > 1:
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)

    for row_idx, fill in (
        (headcode_row, header_fill),
        (header_row, header_fill),
        (facility_row, facility_fill),
    ):
        if not row_idx:
            continue
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row_idx, col_idx)
            if cell.hyperlink:
                cell.font = Font(bold=True, color="0563C1", underline="single")
            else:
                cell.font = bold_font
            cell.fill = fill

    for row_idx in (headcode_row, header_row, facility_row, *(extra_rows or [])):
        if row_idx and max_col >= 2:
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)

    station_band = -1
    for row_idx in range((facility_row or header_row or 0) + 1, max_row + 1):
        first_cell_text = str(ws.cell(row_idx, 1).value or "").strip()
        if first_cell_text:
            station_band += 1
            row_border = station_top_border
        else:
            row_border = border
        fill = band_fills[station_band % len(band_fills)] if station_band >= 0 else band_fills[0]
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.fill = fill
            cell.border = row_border
        if first_cell_text:
            ws.cell(row_idx, 1).font = bold_font

    ws.freeze_panes = ws.cell((facility_row or header_row or 1) + 1, 3)
    _fit_columns(ws, ignore_rows=title_rows)
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 6


def _format_lookup_sheet(ws):
    from openpyxl.styles import Border, Font, PatternFill, Side

    thin_side = Side(style="thin", color="B8B8B8")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    ws.freeze_panes = "A2"
    _fit_columns(ws, max_width=48)


def build_timetable_xlsx(tables, station_codes=None, toc_codes=None):
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)

    used_names = set()
    for idx, table in enumerate(tables or [], start=1):
        base_name = _sheet_name(table.get("sheetName") or table.get("title"), f"Sheet {idx}")
        name = base_name
        suffix = 2
        while name in used_names:
            suffix_text = f" {suffix}"
            name = f"{base_name[:31 - len(suffix_text)]}{suffix_text}"
            suffix += 1
        used_names.add(name)

        ws = wb.create_sheet(name)
        title_rows = []
        if table.get("title"):
            ws.append([str(table.get("title"))])
            title_rows.append(1)
        if table.get("dateLabel"):
            ws.append([str(table.get("dateLabel"))])
            title_rows.append(ws.max_row)
        if table.get("title") or table.get("dateLabel"):
            ws.append([])
        headcode_row = None
        if table.get("headcodes"):
            _append_row(ws, table.get("headcodes", []))
            headcode_row = ws.max_row
        header_row = None
        if table.get("headers"):
            _append_row(ws, table.get("headers", []))
            header_row = ws.max_row
        facility_row = ws.max_row + 1 if table.get("rows") else None
        extra_rows = _append_rows(ws, table.get("rows", []))
        _format_timetable_sheet(
            ws,
            title_rows,
            headcode_row,
            header_row,
            facility_row,
            extra_rows=extra_rows,
        )

    codes_ws = wb.create_sheet("Station codes")
    codes_ws.append(["Code", "Station"])
    for entry in station_codes or []:
        if isinstance(entry, dict):
            code = str(entry.get("code") or "").strip()
            name = str(entry.get("name") or "").strip()
        else:
            code = str(entry[0] if len(entry) > 0 else "").strip()
            name = str(entry[1] if len(entry) > 1 else "").strip()
        if code or name:
            codes_ws.append([code, name])
    _format_lookup_sheet(codes_ws)

    toc_ws = wb.create_sheet("TOC codes")
    toc_ws.append(["Code", "Operator"])
    for entry in toc_codes or []:
        if isinstance(entry, dict):
            code = str(entry.get("code") or "").strip()
            name = str(entry.get("name") or "").strip()
        else:
            code = str(entry[0] if len(entry) > 0 else "").strip()
            name = str(entry[1] if len(entry) > 1 else "").strip()
        if code or name:
            toc_ws.append([code, name])
    _format_lookup_sheet(toc_ws)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
